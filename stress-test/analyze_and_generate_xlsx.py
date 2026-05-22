#!/usr/bin/env python3
"""Read test_battery.csv, generate stress_analysis.xlsx with 4 sheets."""

import csv
import os
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, ScatterChart, Series
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl")
    raise SystemExit(1)

CSV_PATH = os.path.join(os.path.dirname(__file__), "..", "test_battery.csv")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "stress_analysis.xlsx")

rows = []
with open(CSV_PATH, newline="") as f:
    reader = csv.DictReader(f)
    for r in reader:
        r["size"] = int(r["size"])
        r["thickness"] = int(r["thickness"])
        r["blocks"] = int(r["blocks"])
        r["hull_blocks"] = int(r["hull_blocks"])
        r["solve_time_ms"] = float(r["solve_time_ms"])
        r["avg_stress_pct"] = float(r["avg_stress_pct"])
        r["max_stress_pct"] = float(r["max_stress_pct"])
        r["min_crush_depth_blocks"] = float(r["min_crush_depth_blocks"])
        r["hull_ratio"] = float(r["hull_ratio"])
        rows.append(r)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
NARROW = Alignment(horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

def write_sheet(ws, headers, data_rows, col_widths=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = NARROW
        c.border = THIN_BORDER
    for ri, d in enumerate(data_rows, 2):
        for ci, v in enumerate(d, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = NARROW
            c.border = THIN_BORDER
            if col_widths and ci <= len(col_widths):
                ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.auto_filter.ref = ws.dimensions

wb = Workbook()

# ── Sheet 1: hull-fix impact ──────────────────────────────────
ws1 = wb.active
ws1.title = "hull-fix impact"
hdr1 = ["shape", "size", "material", "thickness", "internal_struct",
        "blocks", "hull_blocks", "hull_ratio"]
data1 = []
for r in rows:
    data1.append([r["shape"], r["size"], r["material"], r["thickness"],
                   r["internal_struct"], r["blocks"], r["hull_blocks"],
                   r["hull_ratio"]])
write_sheet(ws1, hdr1, data1, col_widths=[14, 6, 10, 10, 16, 8, 12, 10])

# ── Sheet 2: material predictions ──────────────────────────────
ws2 = wb.create_sheet("material predictions")
hdr2 = ["material", "internal_struct", "shape", "size", "thickness",
        "crush_depth", "avg_stress_pct", "max_stress_pct"]
data2 = []
for r in rows:
    data2.append([r["material"], r["internal_struct"], r["shape"],
                   r["size"], r["thickness"],
                   r["min_crush_depth_blocks"], r["avg_stress_pct"],
                   r["max_stress_pct"]])
write_sheet(ws2, hdr2, data2, col_widths=[10, 16, 14, 6, 10, 12, 14, 14])

# Chart: crush depth by material for solid boxes
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Crush Depth by Material (box, solid, T=1)"
chart1.x_axis.title = "Material"
chart1.y_axis.title = "Crush Depth (blocks)"
materials = ["oak", "iron", "diamond"]
crush_vals = []
for m in materials:
    for r in rows:
        if r["shape"] == "box" and r["material"] == m and r["thickness"] == 1 and r["internal_struct"] == "solid":
            crush_vals.append(r["min_crush_depth_blocks"])
            break
if crush_vals:
    ws2.cell(row=1, column=10, value="material")
    ws2.cell(row=2, column=10, value="oak")
    ws2.cell(row=3, column=10, value="iron")
    ws2.cell(row=4, column=10, value="diamond")
    ws2.cell(row=1, column=11, value="crush")
    for i, v in enumerate(crush_vals, 2):
        ws2.cell(row=i, column=11, value=v)
    chart1.add_data(Reference(ws2, min_col=11, min_row=1, max_row=4))
    chart1.set_categories(Reference(ws2, min_col=10, min_row=2, max_row=4))
    chart1.shape = 4
    ws2.add_chart(chart1, "M2")

# ── Sheet 3: thickness scaling ─────────────────────────────────
ws3 = wb.create_sheet("thickness scaling")
hdr3 = ["shape", "size", "material", "internal_struct", "thickness",
        "blocks", "hull_blocks", "hull_ratio", "crush_depth"]
data3 = []
for r in rows:
    data3.append([r["shape"], r["size"], r["material"], r["internal_struct"],
                   r["thickness"], r["blocks"], r["hull_blocks"],
                   r["hull_ratio"], r["min_crush_depth_blocks"]])
write_sheet(ws3, hdr3, data3, col_widths=[14, 6, 10, 16, 10, 8, 12, 10, 12])

# Crush depth vs thickness for box oak hollow
chart2 = ScatterChart()
chart2.title = "Crush Depth vs Thickness (box oak hollow)"
chart2.x_axis.title = "Thickness"
chart2.y_axis.title = "Crush Depth (blocks)"
if len([r for r in rows if r["shape"] == "box" and r["material"] == "oak" and r["internal_struct"] == "hollow"]) >= 2:
    col_start = 11
    ws3.cell(row=1, column=col_start, value="thickness")
    ws3.cell(row=1, column=col_start + 1, value="crush")
    idx = 2
    for r in rows:
        if r["shape"] == "box" and r["material"] == "oak" and r["internal_struct"] == "hollow":
            ws3.cell(row=idx, column=col_start, value=r["thickness"])
            ws3.cell(row=idx, column=col_start + 1, value=r["min_crush_depth_blocks"])
            idx += 1
    xvals = Reference(ws3, min_col=col_start, min_row=2, max_row=idx - 1)
    yvals = Reference(ws3, min_col=col_start + 1, min_row=1, max_row=idx - 1)
    chart2.add_data(yvals)
    chart2.set_categories(xvals)
    ws3.add_chart(chart2, "K2")

# ── Sheet 4: surface verification ──────────────────────────────
ws4 = wb.create_sheet("surface verification")
hdr4 = ["shape", "size", "material", "thickness", "internal_struct",
        "blocks", "hull_blocks", "hull_ratio", "geom_hull_est",
        "hull_geom_ratio"]
data4 = []
for r in rows:
    s = r["shape"]
    sz = r["size"]
    t = r["thickness"]
    if s == "box":
        inner = max(0, sz - 2 * t)
        geom_hull = sz ** 3 - inner ** 3
    elif s == "sphere":
        # Approximate surface blocks for sphere radius=sz
        vol_ext = int(4 / 3 * 3.14159 * (sz + 0.5) ** 3)
        vol_int = int(4 / 3 * 3.14159 * max(0, sz - t + 0.5) ** 3)
        geom_hull = max(0, vol_ext - vol_int)
    elif s.startswith("ellipsoid"):
        parts = s.split("_")
        if len(parts) == 2:
            a, b, c = 1, 1, 1
        else:
            # Parse like "ellipsoid_211" -> a=2, b=1, c=1
            try:
                nums = parts[1]
                a = int(nums[0]) if len(nums) > 0 else 1
                b = int(nums[1]) if len(nums) > 1 else 1
                c = int(nums[2]) if len(nums) > 2 else 1
            except (IndexError, ValueError):
                a, b, c = 1, 1, 1
        rx, ry, rz = a * sz, b * sz, c * sz
        vol_ext = int(4 / 3 * 3.14159 * (rx + 0.5) * (ry + 0.5) * (rz + 0.5))
        vol_int = int(4 / 3 * 3.14159 * max(0, rx - t + 0.5) * max(0, ry - t + 0.5) * max(0, rz - t + 0.5))
        geom_hull = max(0, vol_ext - vol_int)
    else:
        geom_hull = r["hull_blocks"]

    geom_ratio = r["hull_blocks"] / geom_hull if geom_hull > 0 else 1.0
    data4.append([s, sz, r["material"], t, r["internal_struct"],
                  r["blocks"], r["hull_blocks"], r["hull_ratio"],
                  geom_hull, round(geom_ratio, 4)])
write_sheet(ws4, hdr4, data4, col_widths=[14, 6, 10, 10, 16, 8, 12, 10, 14, 14])

wb.save(XLSX_PATH)
print(f"Written {XLSX_PATH}")
